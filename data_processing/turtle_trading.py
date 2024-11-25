import multiprocessing
import pandas as pd
import numpy as np
import yfinance as yf
import openpyxl
from openpyxl.styles import PatternFill
import mplfinance as mpf
from typing import Optional, List, Dict, Tuple
import os

nifty_500 = ['MMTC', 'DATAPATTNS', 'JYOTHYLAB', 'ADANIGREEN', 'UBL', 'SYNGENE', 'TATAMTRDVR', 'TATAMOTORS', 'NUVAMA', 'HINDPETRO', 'OIL', 'IOC', 'JAIBALAJI', 'RAYMOND', 'PETRONET', 'FORTIS', 'ONGC', 'SUVENPHAR', 'INDIACEM', 'VAIBHAVGBL', 'UNITDSPR', 'CGPOWER', 'BIOCON', 'RAINBOW', 'IDBI', 'SWANENERGY', 'KFINTECH', 'GLS', 'BPCL', 'SBILIFE', 'QUESS', 'JSWINFRA', 'TVSSCS', 'LT', 'MARICO', 'CLEAN', 'LXCHEM', 'RAMCOCEM', 'MCX', 'VTL', 'SANOFI', 'VBL', 'HAVELLS', 'SCHNEIDER', 'ICICIGI', 'BSE', 'GODFRYPHLP', 'GRANULES', 'ABFRL', 'ZEEL', 'INOXWIND', 'SUNPHARMA', 'JUSTDIAL', 'GAIL', 'ZYDUSLIFE', 'SUZLON', 'LINDEINDIA', 'BBTC', 'ADANIENSOL', 'DEEPAKNTR', 'MOTILALOFS', 'LLOYDSME', 'CUMMINSIND', 'TTML', 'APLLTD', 'TIINDIA', 'CCL', 'BRIGADE', 'KOTAKBANK', 'AMBER', 'ABBOTINDIA', 'MANKIND', 'FEDERALBNK', 'KPRMILL', 'ERIS', 'ABCAPITAL', 'BOSCHLTD', 'MHRIL', 'MSUMI', 'POLYCAB', 'CAMPUS', 'INDIGO', 'AFFLE', 'ABB', 'MTARTECH', 'RADICO', 'KEI', 'WIPRO', 'ELGIEQUIP', 'IPCALAB', 'HUDCO', 'UTIAMC', 'GAEL', 'GMMPFAUDLR', 'RENUKA', 'CDSL', 'HDFCLIFE', 'SRF', 'PAYTM', 'SOBHA', 'PEL', 'DCMSHRIRAM', 'ICICIPRULI', 'TIMKEN', 'PAGEIND', 'TATAPOWER', 'SIEMENS', 'TORNTPOWER', 'WHIRLPOOL', 'ASTERDM', 'KAYNES', 'GODREJCP', 'PNB', 'CELLO', 'JUBLINGREA', 'IEX', 'WELSPUNLIV', 'GESHIP', 'ZENSARTECH', 'NAUKRI', 'SONACOMS', 'JUBLPHARMA', 'GSPL', 'JMFINANCIL', 'SUNTV', 'AUROPHARMA', 'BHEL', 'ADANIENT', 'BAJFINANCE', 'LAURUSLABS', 'HDFCBANK', 'PGHH', 'KARURVYSYA', 'GLENMARK', 'RECLTD', 'MOTHERSON', 'TRENT', 'INTELLECT', 'NAM-INDIA', 'JBMA', 'AAVAS', 'USHAMART', 'AJANTPHARM', 'BEL', 'RRKABEL', 'COROMANDEL', 'TCS', 'NETWORK18', 'CENTURYTEX', 'J&KBANK', 'DRREDDY', 'EMAMILTD', 'KANSAINER', 'COALINDIA', 'CYIENT', 'LTTS', 'SUNDRMFAST', 'MGL', 'PNBHOUSING', 'MEDPLUS', 'CHENNPETRO', 'GNFC', 'TECHM', 'PRSMJOHNSN', 'KIMS', 'DIVISLAB', 'DALBHARAT', 'BHARATFORG', 'NATCOPHARM', 'FDC', 'ENDURANCE', 'PFC', 'ADANIPORTS', 'DEEPAKFERT', 'BAJAJ-AUTO', 'ADANIPOWER', 'TATATECH', 'BALKRISIND', 'COLPAL', 'POWERGRID', 'BLUESTARCO', 'AARTIIND', 'APTUS', 'HDFCAMC', 'MFSL', 'CRISIL', 'GRAPHITE', 'THERMAX', 'BERGEPAINT', 'EXIDEIND', 'MUTHOOTFIN', 'CROMPTON', 'HCLTECH', 'SKFINDIA', 'KSB', 'JINDALSAW', 'M&M', 'CSBBANK', 'RHIM', 'ULTRACEMCO', 'LTF', 'IOB', 'VEDL', 'GPIL', 'JBCHEPHARM', 'HAPPYFORGE', 'PIIND', 'TANLA', 'PIDILITIND', 'DMART', 'MASTEK', 'MARUTI', 'CASTROLIND', 'MRF', 'EIHOTEL', 'BRITANNIA', 'COFORGE', 'MEDANTA', 'AEGISLOG', 'METROPOLIS', 'CIPLA', 'SIGNATURE', 'PVRINOX', 'NUVOCO', 'AWL', 'ALKYLAMINE', 'RELIANCE', 'OFSS', 'DIXON', 'LICI', 'TRIDENT', 'TVSMOTOR', 'NTPC', 'BAJAJFINSV', 'GRASIM', 'SYRMA', 'IRCTC', 'ASIANPAINT', 'ARE&M', 'HINDCOPPER', 'HINDUNILVR', 'ENGINERSIN', 'HONASA', 'TRITURBINE', 'ALKEM', 'CENTRALBK', 'REDINGTON', 'HAL', 'TATACHEM', 'INDUSTOWER', 'ASHOKLEY', 'ITI', 'GLAXO', 'GILLETTE', 'SBIN', 'BAYERCROP', 'BHARTIARTL', 'LODHA', 'STARHEALTH', 'UCOBANK', 'SUMICHEM', 'APOLLOTYRE', 'NHPC', 'NAVINFLUOR', 'FSL', 'GODREJIND', 'CANBK', 'EICHERMOT', 'LEMONTREE', 'INFY', 'HEROMOTOCO', 'SAREGAMA', 'BALAMINES', 'CENTURYPLY', 'NIACL', 'ALLCARGO', 'JUBLFOOD', 'LATENTVIEW', 'GMDCLTD', 'HONAUT', 'TMB', 'PRAJIND', 'ATUL', 'BATAINDIA', 'PATANJALI', 'DLF', 'SHREECEM', 'DABUR', 'CARBORUNIV', 'FINCABLES', 'RAJESHEXPO', 'HINDALCO', 'INDIAMART', 'FLUOROCHEM', 'INDIANB', 'TATACONSUM', 'VIPIND', 'CUB', 'NCC', 'PPLPHARMA', 'KPIL', 'HOMEFIRST', 'HEG', 'RTNINDIA', 'MRPL', 'ASTRAL', 'BAJAJHLDNG', 'APOLLOHOSP', 'CESC', 'MAHSEAMLES', 'PNCINFRA', 'AMBUJACEM', 'TATACOMM', 'CHEMPLASTS', 'GODREJPROP', 'KNRCON', 'BALRAMCHIN', 'MPHASIS', 'CIEINDIA', 'GUJGASLTD', 'ITC', 'OBEROIRLTY', 'RBA', 'JSL', 'JSWSTEEL', 'ATGL', 'GLAND', 'TV18BRDCST', 'EQUITASBNK', 'CGCL', 'IBULHSGFIN', 'CHAMBLFERT', 'MAHABANK', 'CONCOR', 'CAMS', 'MAPMYINDIA', 'YESBANK', 'LUPIN', 'MANYAVAR', 'SUNTECK', 'LALPATHLAB', 'FINEORG', 'SPARC', 'AETHER', 'UNOMINDA', 'CAPLIPOINT', 'KEC', 'NATIONALUM', 'CHOLAHLDNG', 'BANKINDIA', 'PRINCEPIPE', 'NMDC', 'METROBRAND', 'GMRINFRA', 'POLYMED', 'FINPIPE', 'POLICYBZR', 'TATAELXSI', 'VGUARD', 'GICRE', 'JKLAKSHMI', 'DELHIVERY', 'SBICARD', 'MINDACORP', 'BANDHANBNK', 'UPL', 'EASEMYTRIP', 'INDUSINDBK', 'NH', 'HFCL', 'IDFC', 'PCBL', 'ZFCVINDIA', 'IDFCFIRSTB', 'KAJARIACER', 'SBFC', 'HSCL', 'UJJIVANSFB', 'BANKBARODA', 'SUNDARMFIN', 'CONCORDBIO', 'CERA', 'OLECTRA', 'NSLNISP', 'GSFC', 'HAPPSTMNDS', 'SWSOLAR', 'LTIM', 'DEVYANI', 'BLS', 'INDHOTEL', 'ACC', 'SHYAMMETL', 'IRB', 'TEJASNET', 'BDL', 'CEATLTD', 'APARINDS', 'SUPREMEIND', 'JKCEMENT', '360ONE', 'ACE', 'BSOFT', 'TATAINVEST', 'GPPL', 'JIOFIN', 'ECLERX', 'SHRIRAMFIN', 'TRIVENI', 'CREDITACC', 'EPL', 'LICHSGFIN', 'TATASTEEL', 'FACT', 'NYKAA', 'NLCINDIA', 'VIJAYA', 'CHOLAFIN', 'PHOENIXLTD', 'SOLARINDS', 'ELECON', 'BIKAJI', 'APLAPOLLO', 'JSWENERGY', 'UNIONBANK', 'SAFARI', 'FIVESTAR', 'BLUEDART', 'TORNTPHARM', 'CRAFTSMAN', 'ANURAS', 'EIDPARRY', 'IDEA', 'SCHAEFFLER', 'ALOKINDS', 'PRESTIGE', 'RKFORGE', 'GRINDWELL', 'RATNAMANI', 'ZOMATO', 'STLTECH', '3MINDIA', 'TITAN', 'MAXHEALTH', 'ICICIBANK', 'ANANDRATHI', 'RAILTEL', 'ASTRAZEN', 'ESCORTS', 'IGL', 'ISEC', 'WESTLIFE', 'IIFL', 'INDIGOPNTS', 'CANFINHOME', 'AIAENG', 'KALYANKJIL', 'ANGELONE', 'DOMS', 'ACI', 'ROUTE', 'PERSISTENT', 'HBLPOWER', 'SJVN', 'NESTLEIND', 'BEML', 'RITES', 'BIRLACORPN', 'SAIL', 'SAPPHIRE', 'ASAHIINDIA', 'POWERINDIA', 'JWL', 'MANAPPURAM', 'RCF', 'RVNL', 'VOLTAS', 'NBCC', 'RBLBANK', 'TITAGARH', 'JINDALSTEL', 'SONATSOFTW', 'COCHINSHIP', 'MAZDOCK', 'M&MFIN', 'IRCON', 'BORORENEW', 'VARROC', 'KPITTECH', 'MAHLIFE', 'KRBL', 'POONAWALLA', 'CHALET', 'HINDZINC', 'AVANTIFEED', 'IRFC', 'AUBANK', 'WELCORP', 'GRSE', 'AXISBANK', 'JKPAPER']


class TurtleTrading:
    def __init__(self, ticker: str, period: str = "1mo", nse: bool = True, rounding: int = 2):
        """
        Turtle Trading logic for fetching data, calculating indicators, and generating signals.
        """
        stock_exchange = ".NS" if nse else ""
        self.ticker = ticker
        self.mod_ticker = ticker + stock_exchange
        self.period = period
        self.rounding = rounding
        self.data: Optional[pd.DataFrame] = None

    def fetch_data(self) -> pd.DataFrame:
        """
        Fetch historical stock data.
        """
        try:
            self.data = yf.download(self.mod_ticker, period=self.period)
            self.data[['Open', 'High', 'Low', 'Close', 'Adj Close']] = self.data[
                ['Open', 'High', 'Low', 'Close', 'Adj Close']
            ].round(self.rounding)
        except Exception as e:
            raise RuntimeError(f"Failed to fetch data for {self.ticker}: {e}")
        return self.data

    def calculate_indicators(self) -> pd.DataFrame:
        """
        Calculate technical indicators like 55-day High, 20-day Low, and ATR.
        """
        if self.data is None:
            raise ValueError("Data not loaded. Call fetch_data first.")

        self.data['55d_High'] = self.data['High'].rolling(window=55).max()
        self.data['20d_Low'] = self.data['Low'].rolling(window=20).min()
        true_range = pd.concat([
            self.data['High'] - self.data['Low'],
            (self.data['High'] - self.data['Close'].shift(1)).abs(),
            (self.data['Close'].shift(1) - self.data['Low']).abs()
        ], axis=1).max(axis=1)
        self.data['True_Range'] = true_range
        self.data['Avg_True_Range(N)'] = true_range.rolling(
            window=20).mean().round(self.rounding)
        return self.data

    def generate_signals(self) -> pd.DataFrame:
        """
        Generate buy/sell signals based on Turtle Trading strategy.
        """
        if self.data is None:
            raise ValueError("Data not loaded. Call fetch_data first.")

        self.data['Buy_Signal'] = self.data['High'] > self.data['55d_High'].shift(
            1)
        self.data['Sell_Signal'] = self.data['Low'] < self.data['20d_Low'].shift(
            1)
        return self.data

    def save_to_excel(self, filename: str):
        """
        Save data to Excel with conditional formatting for Buy/Sell signals.
        """
        if self.data is None:
            raise ValueError("Data not loaded. Call fetch_data first.")

            # Check if the file exists; create it if it doesn't
        if not os.path.exists(filename):
            # Create an empty workbook and save it
            workbook = openpyxl.Workbook()
            workbook.save(filename)

        with pd.ExcelWriter(filename, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
            self.data.to_excel(writer, sheet_name=self.ticker)

    def generate_format(self, filename: str):
        """
        Apply conditional formatting for Buy_Signal and Sell_Signal in the Excel file and freeze panes.
        """
        if self.data is None:
            raise ValueError("Data not loaded. Call fetch_data first.")

        # Load the workbook and worksheet
        try:
            workbook = openpyxl.load_workbook(filename)
        except FileNotFoundError:
            raise FileNotFoundError(
                f"The file {filename} does not exist. Save data first.")

        if self.ticker not in workbook.sheetnames:
            raise ValueError(f"Sheet for ticker {self.ticker} not found in {
                             filename}. Ensure data is saved.")

        worksheet = workbook[self.ticker]

        # Define fill styles for Buy_Signal and Sell_Signal
        green_fill = PatternFill(start_color='C6EFCE',
                                 end_color='C6EFCE', fill_type='solid')
        red_fill = PatternFill(start_color='FFC7CE',
                               end_color='FFC7CE', fill_type='solid')

        # Locate column indices for Buy_Signal and Sell_Signal
        buy_signal_col = self.data.columns.get_loc('Buy_Signal') + 1
        sell_signal_col = self.data.columns.get_loc('Sell_Signal') + 1

        # Apply conditional formatting row by row
        for row in range(2, len(self.data) + 2):  # Skip header row
            buy_signal = worksheet.cell(row, buy_signal_col + 1).value
            sell_signal = worksheet.cell(row, sell_signal_col + 1).value

            if buy_signal:  # Apply green fill if Buy_Signal is True
                for cell in worksheet[row]:
                    cell.fill = green_fill
            if sell_signal:  # Apply red fill if Sell_Signal is True
                for cell in worksheet[row]:
                    cell.fill = red_fill

        # Freeze the first row and first column
        worksheet.freeze_panes = 'B2'

        # Save changes to the workbook
        workbook.save(filename)
        print(f"Formatting applied and saved to {filename}.")

    def plot_candlestick(self):
        """
        Plot a candlestick chart.
        """
        if self.data is None:
            raise ValueError("Data not loaded. Call fetch_data first.")
        mpf.plot(self.data, type="candle", volume=False,
                 style="charles", title=self.ticker)


def process_tickers(tickers: list, filename: str, days: int = 1, period: str = "1y"):
    """
    Process multiple tickers and save data for those meeting the buy signal.
    """
    selective_tickers = {}
    for ticker in tickers:
        trader = TurtleTrading(ticker, period=period)
        trader.fetch_data()
        trader.calculate_indicators()
        data = trader.generate_signals()
        if data['Buy_Signal'].tail(days).any():
            trader.save_to_excel(filename)
            trader.generate_format(filename)
            last_row = data.iloc[-1]
            atr = last_row['Avg_True_Range(N)']
            close_price = last_row['Close']
            selective_tickers[ticker] = round((atr / close_price) * 100, 4)
    return dict(sorted(selective_tickers.items(), key=lambda x: x[1], reverse=True))

# Example usage
if __name__ == "__main1__":
    from nifty_50_tickers import nifty_500

    tickers = nifty_500
    filename = "Turtle1_Trading.xlsx"
    top_tickers = process_tickers(tickers, filename, days=1, period="1y")
    print(top_tickers)

def process_single_ticker(ticker: str, period: str, filename: str, days: int) -> Tuple[str, float]:
    """
    Process a single ticker to fetch data, calculate indicators, and check buy signals.
    Returns a tuple of the ticker and its ATR percentage if it meets the buy signal criteria.
    """
    try:
        # Initialize TurtleTrading for the ticker
        trader = TurtleTrading(ticker, period=period)
        trader.fetch_data()
        trader.calculate_indicators()
        data = trader.generate_signals()

        # Check for buy signals
        if data['Buy_Signal'].tail(days).any():
            # Save data to Excel
            trader.save_to_excel(filename)
            trader.generate_format(filename)
            last_row = data.iloc[-1]
            atr = last_row['Avg_True_Range(N)']
            close_price = last_row['Close']
            atr_percentage = round((atr / close_price) * 100, 4)
            return ticker, atr_percentage

    except Exception as e:
        print(f"Error processing {ticker}: {e}")
    return ticker, None


def process_tickers_multiprocessing(tickers: List[str], filename: str, days: int = 1, period: str = "1y") -> Dict[str, float]:
    """
    Process multiple tickers in parallel using multiprocessing.
    Returns a dictionary of tickers and their ATR percentages for those meeting the buy signal criteria.
    """

    # Create a partial function with fixed arguments
    with multiprocessing.Pool(processes=multiprocessing.cpu_count()) as pool:
        results = pool.starmap(
            process_single_ticker, [(ticker, period, filename, days) for ticker in tickers]
        )
    # Filter out None results 5pand sort by ATR percentage
    filtered_results = {ticker: atr for ticker,
                        atr in results if atr is not None}
    return dict(sorted(filtered_results.items(), key=lambda x: x[1], reverse=True))


if __name__ == "__main__":
    # from .. nifty_50_tickers import nifty_500

    tickers = nifty_500
    filename = "Turtle.xlsx"
    top_tickers = process_tickers_multiprocessing(
        tickers, filename, days=1, period="1y")
    print(top_tickers)

